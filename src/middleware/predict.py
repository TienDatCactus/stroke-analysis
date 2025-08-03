import os
import sys
import json
import joblib
import logging
import traceback
import warnings
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Set up logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('stroke_prediction')


def validate_excel_file(path):
    """Validate if Excel file is readable and has at least one worksheet."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    try:
        wb = load_workbook(path, read_only=True)
        if not wb.sheetnames:
            raise ValueError("Excel file contains no sheets.")
        wb.close()
    except Exception as e:
        raise ValueError(f"Invalid Excel file format: {str(e)}")


def clean_and_prepare_data(df):
    """Clean, normalize, and extract numeric features from DataFrame."""
    # Replace invalid strings with zero
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=FutureWarning)
        df.replace(['#NULL!', 'NULL', 'NA'], 0, inplace=True)

    # Drop unnecessary features
    columns_to_drop = [
        "Dizzy", "hearing loss", "Horizontal diplopia", "Vertical diplopia",
        "Other double vision", "Headache", "Loss of vision", "Memory disorder",
        "corresponding intracranial occlusion site ICA",
        "corresponding intracranial occlusion site MCA",
        "corresponding intracranial occlusion site BA",
        "corresponding intracranial occlusion site ACA",
        "corresponding intracranial occlusion site P2",
        "corresponding intracranial occlusion site PCA",
        "anticoagulant vitamin K", "apixaban", "dabigatran", "rivaroxaban",
        "Statin intensity level", "diabetes medication treatment",
        "hypertension medication treatment", "Days in hospital",
        "stroke recurrence within 30 days", "stroke recurrence within 90 days"
    ]
    columns_found = [col for col in columns_to_drop if col in df.columns]
    if len(columns_found) != len(columns_to_drop):
        logger.warning(f"Some columns to drop weren't found in the data. Expected: {len(columns_to_drop)}, Found: {len(columns_found)}")

    df = df.drop(columns=columns_found, errors='ignore')
    df = df.fillna(0)
    df.columns = df.columns.str.lower()

    # Drop identifier columns
    id_columns = ['end', 'no.']
    df = df.drop(columns=[col for col in id_columns if col in df.columns], errors='ignore')

    # Select numerical data only
    return df.select_dtypes(include=['number'])


def load_model(model_path):
    """Load trained machine learning model from file."""
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model file not found: {model_path}")
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning)
        return joblib.load(model_path)


def predict_new_data(link_new_data, model_path):
    """Main prediction logic."""
    try:
        logger.info(f"Validating Excel file: {link_new_data}")
        validate_excel_file(link_new_data)

        logger.info("Reading Excel data")
        new_data = pd.read_excel(link_new_data, sheet_name=0)
        logger.info(f"Data loaded: {new_data.shape[0]} rows, {new_data.shape[1]} columns")

        logger.info("Cleaning and preparing data")
        data = clean_and_prepare_data(new_data)
        logger.info(f"Using {data.shape[1]} features for prediction: {', '.join(data.columns)}")

        logger.info(f"Loading model: {model_path}")
        model = load_model(model_path)
        logger.info("Model loaded successfully")

        X_new = data.to_numpy()
        predictions = model.predict(X_new)
        predictions = np.where(predictions == 0, 'Stroke', 'No Stroke')

        results = [{"index": idx + 1, "prediction": pred} for idx, pred in enumerate(predictions)]
        logger.info(f"Prediction completed: {len(results)} records")

        return {
            "success": True,
            "predictions": predictions.tolist(),
            "results": results
        }

    except Exception as e:
        logger.error(f"Error in prediction: {str(e)}")
        logger.error(traceback.format_exc())

        # Identify specific error codes
        error_text = str(e).lower()
        error_code = "UNKNOWN_ERROR"
        missing_columns, data_type_issues = None, None

        if any(k in error_text for k in ["keyerror", "missing", "no column"]):
            error_code = "MISSING_COLUMNS"
            import re
            column_matches = re.findall(r"['\"](.*?)['\"]", str(e))
            if column_matches:
                missing_columns = column_matches
        elif any(k in error_text for k in ["could not convert", "invalid literal", "dtype"]):
            error_code = "INVALID_DATA_TYPE"
        elif any(k in error_text for k in ["file format", "read_excel", "parse"]):
            error_code = "FILE_FORMAT_ERROR"
        elif any(k in error_text for k in ["model", "predict", "sklearn"]):
            error_code = "MODEL_ERROR"

        return {
            "success": False,
            "error": str(e),
            "errorCode": error_code,
            "missingColumns": missing_columns,
            "dataTypeIssues": data_type_issues,
            "predictions": [],
            "results": []
        }


def predict_new_data_cli(link_new_data, model_path):
    """CLI wrapper for predict_new_data that outputs JSON to stdout."""
    result = predict_new_data(link_new_data, model_path)
    print(json.dumps(result, indent=2))
    return result


if __name__ == "__main__":
    import platform
    logger.info(f"Python version: {platform.python_version()}")
    logger.info(f"Python implementation: {platform.python_implementation()}")

    try:
        import importlib.metadata as metadata
        installed_packages = [f"{dist.metadata['Name']}=={dist.version}" for dist in metadata.distributions()]
        logger.info(f"Installed packages (partial): {', '.join(installed_packages[:5])}...")
    except Exception as e:
        logger.warning(f"Couldn't list installed packages: {str(e)}")

    if len(sys.argv) < 3:
        print("Usage: python predict.py <excel_file_path> <model_path>")
        sys.exit(1)

    result = predict_new_data_cli(sys.argv[1], sys.argv[2])
    sys.exit(0 if result["success"] else 1)
